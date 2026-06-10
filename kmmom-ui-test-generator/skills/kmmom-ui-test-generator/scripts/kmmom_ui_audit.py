# -*- coding: utf-8 -*-
from __future__ import annotations

import argparse
import base64
import datetime as dt
import json
import os
import re
import socket
import struct
import time
import urllib.parse
import urllib.request
import urllib.error
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


DEFAULT_BASE_URL = "http://192.168.30.69:40000"
DEFAULT_DEBUG_URL = "http://127.0.0.1:9223/json/list"
RISK_PATTERNS = [
    "导出",
    "下载",
    "保存",
    "提交",
    "确定",
    "确 定",
    "删除",
    "初始化",
    "发布",
    "下发",
    "通过",
    "驳回",
    "发送消息",
]
OUTPUT_HEADERS = [
    "序号",
    "模块",
    "菜单路径",
    "所属页面",
    "路由地址",
    "区域",
    "按钮/控件",
    "控件类型",
    "点击前置",
    "预期结果",
    "实测状态",
    "展开/弹窗/选项内容",
    "异常/弹窗提示",
    "是否执行点击",
    "风险说明",
]


class CDP:
    def __init__(self, ws_url: str):
        rest = ws_url[5:]
        hostport, path = rest.split("/", 1)
        path = "/" + path
        if ":" in hostport:
            host, port_s = hostport.split(":", 1)
            port = int(port_s)
        else:
            host, port = hostport, 80
        self.sock = socket.create_connection((host, port), timeout=8)
        key = base64.b64encode(os.urandom(16)).decode("ascii")
        request = (
            f"GET {path} HTTP/1.1\r\n"
            f"Host: {hostport}\r\n"
            "Upgrade: websocket\r\n"
            "Connection: Upgrade\r\n"
            f"Sec-WebSocket-Key: {key}\r\n"
            "Sec-WebSocket-Version: 13\r\n\r\n"
        )
        self.sock.sendall(request.encode("ascii"))
        response = b""
        while b"\r\n\r\n" not in response:
            chunk = self.sock.recv(4096)
            if not chunk:
                raise RuntimeError("CDP websocket handshake failed")
            response += chunk
        if b" 101 " not in response.split(b"\r\n", 1)[0]:
            raise RuntimeError(response.decode("utf-8", errors="replace"))
        self.buf = b""
        self.next_id = 1

    def _send_frame(self, payload: str) -> None:
        data = payload.encode("utf-8")
        header = bytearray([0x81])
        n = len(data)
        if n < 126:
            header.append(0x80 | n)
        elif n < 65536:
            header.append(0x80 | 126)
            header.extend(struct.pack("!H", n))
        else:
            header.append(0x80 | 127)
            header.extend(struct.pack("!Q", n))
        mask = os.urandom(4)
        header.extend(mask)
        self.sock.sendall(bytes(header) + bytes(b ^ mask[i % 4] for i, b in enumerate(data)))

    def _recv_frame(self) -> str:
        def need(n: int) -> None:
            while len(self.buf) < n:
                chunk = self.sock.recv(65536)
                if not chunk:
                    raise EOFError("CDP websocket closed")
                self.buf += chunk

        need(2)
        b1, b2 = self.buf[0], self.buf[1]
        self.buf = self.buf[2:]
        opcode = b1 & 0x0F
        ln = b2 & 0x7F
        if ln == 126:
            need(2)
            ln = struct.unpack("!H", self.buf[:2])[0]
            self.buf = self.buf[2:]
        elif ln == 127:
            need(8)
            ln = struct.unpack("!Q", self.buf[:8])[0]
            self.buf = self.buf[8:]
        mask = None
        if b2 & 0x80:
            need(4)
            mask = self.buf[:4]
            self.buf = self.buf[4:]
        need(ln)
        data = self.buf[:ln]
        self.buf = self.buf[ln:]
        if mask:
            data = bytes(b ^ mask[i % 4] for i, b in enumerate(data))
        if opcode == 8:
            raise EOFError("CDP websocket closed")
        return data.decode("utf-8", errors="replace")

    def call(self, method: str, params: dict | None = None, timeout: float = 12) -> dict:
        msg_id = self.next_id
        self.next_id += 1
        msg = {"id": msg_id, "method": method}
        if params is not None:
            msg["params"] = params
        self._send_frame(json.dumps(msg, ensure_ascii=False))
        end = time.time() + timeout
        while time.time() < end:
            self.sock.settimeout(max(0.1, end - time.time()))
            try:
                obj = json.loads(self._recv_frame())
            except socket.timeout:
                continue
            except Exception:
                continue
            if obj.get("id") == msg_id:
                if "error" in obj:
                    raise RuntimeError(obj["error"])
                return obj.get("result", {})
        raise TimeoutError(method)

    def eval(self, expression: str, timeout: float = 12):
        result = self.call(
            "Runtime.evaluate",
            {
                "expression": expression,
                "awaitPromise": True,
                "returnByValue": True,
                "userGesture": True,
                "timeout": int(timeout * 1000),
            },
            timeout + 2,
        )
        return result.get("result", {}).get("value")

    def navigate(self, url: str) -> None:
        self.call("Page.navigate", {"url": url}, timeout=8)

    def esc(self) -> None:
        for typ in ["keyDown", "keyUp"]:
            self.call(
                "Input.dispatchKeyEvent",
                {"type": typ, "key": "Escape", "code": "Escape", "windowsVirtualKeyCode": 27, "nativeVirtualKeyCode": 27},
                timeout=2,
            )


def js_string(value: str) -> str:
    return json.dumps(value, ensure_ascii=False)


def text(value) -> str:
    if value is None:
        return ""
    return str(value).replace("\r\n", "\n").replace("\r", "\n").strip()


def clipped(value, limit: int = 900) -> str:
    value = re.sub(r"\s+", " ", text(value))
    if len(value) <= limit:
        return value
    return value[: limit - 12] + "...[截断]"


def is_risky(name: str) -> bool:
    name = text(name)
    return any(p in name for p in RISK_PATTERNS)


def page_url(base_url: str, route: str) -> str:
    base = base_url.rstrip("/")
    route = route if route.startswith("/") else "/" + route
    return base + "/#" + route


def connect(debug_url: str, base_url: str) -> CDP:
    try:
        targets = json.loads(urllib.request.urlopen(debug_url, timeout=8).read().decode("utf-8"))
    except urllib.error.URLError as exc:
        raise RuntimeError(
            f"Cannot connect to Chrome debug endpoint: {debug_url}. "
            "Launch Chrome with scripts/launch_chrome.ps1 or confirm the configured remote debugging port is open."
        ) from exc
    parsed = urllib.parse.urlparse(base_url)
    host = parsed.hostname or ""
    page = next((t for t in targets if t.get("type") == "page" and host in t.get("url", "")), None)
    page = page or next((t for t in targets if t.get("type") == "page"), None)
    if not page:
        raise RuntimeError(f"No Chrome page target found from {debug_url}. Launch Chrome with remote debugging first.")
    cdp = CDP(page["webSocketDebuggerUrl"])
    cdp.call("Runtime.enable")
    cdp.call("Page.enable")
    return cdp


def wait_ready(cdp: CDP, seconds: float = 2.0) -> None:
    time.sleep(seconds)
    for _ in range(8):
        try:
            state = cdp.eval("document.readyState", timeout=3)
            if state in {"interactive", "complete"}:
                time.sleep(0.6)
                return
        except Exception:
            pass
        time.sleep(0.5)


def login_if_needed(cdp: CDP, username: str, password: str) -> bool:
    if not username or not password:
        return False
    script = f"""
(async () => {{
  const visible = el => {{
    if (!el) return false;
    const s = getComputedStyle(el);
    const r = el.getBoundingClientRect();
    return s.display !== 'none' && s.visibility !== 'hidden' && r.width > 0 && r.height > 0;
  }};
  const pass = [...document.querySelectorAll('input[type="password"]')].find(visible);
  if (!pass) return false;
  const inputs = [...document.querySelectorAll('input')].filter(visible);
  const user = inputs.find(i => i !== pass && (i.type === 'text' || !i.type || i.getAttribute('autocomplete') === 'username')) || inputs[0];
  const setVal = (el, val) => {{
    const proto = Object.getPrototypeOf(el);
    const desc = Object.getOwnPropertyDescriptor(proto, 'value');
    if (desc && desc.set) desc.set.call(el, val); else el.value = val;
    el.dispatchEvent(new Event('input', {{bubbles:true}}));
    el.dispatchEvent(new Event('change', {{bubbles:true}}));
  }};
  if (user) setVal(user, {js_string(username)});
  setVal(pass, {js_string(password)});
  await new Promise(r => setTimeout(r, 200));
  const buttons = [...document.querySelectorAll('button,[role="button"]')].filter(visible);
  const login = buttons.find(b => /登录|登 录|Login|Sign in/i.test(b.innerText || b.textContent || '')) || buttons.find(b => b.type === 'submit') || buttons[0];
  if (login) login.click();
  return true;
}})()
"""
    try:
        result = bool(cdp.eval(script, timeout=8))
        if result:
            wait_ready(cdp, 4)
        return result
    except Exception:
        return False


def get_module_pages(cdp: CDP, module: str) -> list[dict]:
    script = f"""
(() => {{
  const keys = ['km-dynamicMenu', 'km-route-guard-menu', 'menu', 'menus'];
  let menu = [];
  for (const key of keys) {{
    try {{
      const raw = localStorage.getItem(key);
      if (!raw) continue;
      const parsed = JSON.parse(raw);
      if (Array.isArray(parsed)) {{ menu = parsed; break; }}
      if (Array.isArray(parsed?.data)) {{ menu = parsed.data; break; }}
    }} catch (e) {{}}
  }}
  const out = [];
  const walk = (nodes, ancestors=[]) => {{
    (nodes || []).forEach((n, idx) => {{
      const name = n.name || n.title || n.meta?.title || n.code || n.path || '';
      const path = n.path || n.linkUrl || n.url || '';
      const children = n.children || n.routes || [];
      const chain = ancestors.concat([{{name, path}}]);
      out.push({{
        id: n.id || n.code || '',
        name,
        path,
        module: chain[0]?.name || '',
        fullPathText: chain.map(x => x.name).filter(Boolean).join(' > '),
        hasChildren: children.length > 0,
        visibleFlag: n.visibleFlag !== false,
        hideInMenu: !!n.hideInMenu,
        level: chain.length,
        sortIndex: n.index ?? n.sort ?? idx
      }});
      walk(children, chain);
    }});
  }};
  walk(menu);
  return out
    .filter(x => x.module === {js_string(module)} && !x.hasChildren && x.path && x.path.startsWith('/') && x.visibleFlag && !x.hideInMenu)
    .sort((a,b) => a.fullPathText.localeCompare(b.fullPathText, 'zh-Hans-CN'));
}})()
"""
    pages = cdp.eval(script, timeout=20) or []
    return pages


HELPER_JS = r"""
(() => {
  const h = {};
  h.visible = el => {
    if (!el) return false;
    const s = getComputedStyle(el);
    const r = el.getBoundingClientRect();
    return s.display !== 'none' && s.visibility !== 'hidden' && r.width > 0 && r.height > 0 &&
      r.bottom > 0 && r.right > 0 && r.top < innerHeight && r.left < innerWidth;
  };
  h.text = el => (el?.innerText || el?.textContent || el?.getAttribute?.('title') || el?.getAttribute?.('aria-label') || el?.getAttribute?.('placeholder') || '').replace(/\s+/g,' ').trim();
  h.rect = el => {
    const r = el.getBoundingClientRect();
    return {x:Math.round(r.x), y:Math.round(r.y), w:Math.round(r.width), h:Math.round(r.height), cx:Math.round(r.x + r.width/2), cy:Math.round(r.y + r.height/2)};
  };
  h.kind = el => {
    const tag = el.tagName.toLowerCase();
    const cls = el.className?.toString?.() || '';
    const role = el.getAttribute('role') || '';
    const type = el.getAttribute('type') || '';
    if (tag === 'button' || role === 'button' || cls.includes('ant-btn')) return '按钮';
    if (tag === 'input' && ['checkbox'].includes(type)) return '复选框';
    if (tag === 'input' && ['radio'].includes(type)) return '单选';
    if (tag === 'input' || tag === 'textarea') {
      if (cls.includes('ant-picker') || el.closest('.ant-picker')) return '日期/时间选择器';
      return '输入框';
    }
    if (cls.includes('ant-select') || el.closest('.ant-select')) return '下拉/选择器';
    if (cls.includes('ant-picker') || el.closest('.ant-picker')) return '日期/时间选择器';
    if (cls.includes('ant-checkbox')) return '复选框';
    if (cls.includes('ant-radio')) return '单选';
    if (tag === 'th') return '表头/排序字段';
    if (cls.includes('ant-tabs-tab')) return '页签';
    return '控件';
  };
  h.area = el => {
    if (el.closest('.ant-modal,.ant-drawer')) return '弹窗/抽屉';
    if (el.closest('.ant-table')) return el.tagName.toLowerCase() === 'th' ? '表格区' : '表格区';
    if (el.closest('.ant-tabs-nav')) return '页面内容区';
    const txt = h.text(el);
    const y = el.getBoundingClientRect().top;
    const cls = (el.closest('[class]')?.className || '').toString().toLowerCase();
    if (/query|search|filter|form|condition/.test(cls) || /查询|重置|展开|收起/.test(txt) || y < 260) return '筛选区';
    if (el.closest('.ant-form')) return '页面条件区';
    if (el.closest('.ant-btn-group') || y < 360) return '工具栏';
    if (el.closest('[class*="float"],[class*="assistant"],[class*="dev-tools"]')) return '公共浮动区';
    return '页面内容区';
  };
  h.name = el => {
    let name = h.text(el);
    if (!name && el.matches('input,textarea')) name = el.getAttribute('placeholder') || el.getAttribute('value') || '';
    if (!name && el.closest('.ant-form-item')) {
      const label = el.closest('.ant-form-item').querySelector('.ant-form-item-label,label');
      name = h.text(label);
    }
    if (!name && el.closest('th')) name = h.text(el.closest('th'));
    if (!name) name = h.kind(el);
    return name || '控件';
  };
  h.controls = () => {
    const selectors = [
      'button',
      '[role="button"]',
      'input',
      'textarea',
      '.ant-select',
      '.ant-picker',
      '.ant-checkbox-wrapper',
      '.ant-radio-wrapper',
      '.ant-tabs-tab',
      '.ant-table th',
      '.dev-tools-component-info-button',
      '.dev-tools-engine-debug-button',
      '.float-assistant-button'
    ].join(',');
    const all = [...document.querySelectorAll(selectors)].filter(h.visible);
    const seen = new Set();
    return all.map(el => {
      const r = h.rect(el);
      const name = h.name(el);
      const kind = h.kind(el);
      const area = h.area(el);
      const key = [area, kind, name, r.x, r.y, r.w, r.h].join('|');
      if (seen.has(key)) return null;
      seen.add(key);
      return {
        area,
        name,
        kind,
        disabled: !!(el.disabled || el.getAttribute('aria-disabled') === 'true' || el.className?.toString?.().includes('disabled')),
        rect: r,
        className: el.className?.toString?.() || '',
        tagName: el.tagName
      };
    }).filter(Boolean);
  };
  window.__kmmomAudit = h;
  return true;
})()
"""


def install_helper(cdp: CDP) -> None:
    cdp.eval(HELPER_JS, timeout=10)


def expand_filters(cdp: CDP) -> None:
    script = r"""
(async () => {
  const visible = el => {
    if (!el) return false;
    const s = getComputedStyle(el);
    const r = el.getBoundingClientRect();
    return s.display !== 'none' && s.visibility !== 'hidden' && r.width > 0 && r.height > 0;
  };
  const buttons = [...document.querySelectorAll('button,[role="button"]')].filter(visible);
  const targets = buttons.filter(b => /展开|更多|高级查询/.test((b.innerText || b.textContent || '').replace(/\s+/g,'')));
  for (const b of targets.slice(0, 3)) {
    b.click();
    await new Promise(r => setTimeout(r, 350));
  }
  return targets.length;
})()
"""
    try:
        cdp.eval(script, timeout=6)
    except Exception:
        pass


def collect_controls(cdp: CDP) -> list[dict]:
    try:
        controls = cdp.eval("window.__kmmomAudit ? window.__kmmomAudit.controls() : []", timeout=12)
        return controls or []
    except Exception:
        return []


def interact(cdp: CDP, control: dict) -> dict:
    r = control.get("rect") or {}
    x, y = int(r.get("cx", 0)), int(r.get("cy", 0))
    kind = text(control.get("kind"))
    script = f"""
(async () => {{
  const x = {x}, y = {y};
  const visible = el => {{
    if (!el) return false;
    const s = getComputedStyle(el);
    const r = el.getBoundingClientRect();
    return s.display !== 'none' && s.visibility !== 'hidden' && r.width > 0 && r.height > 0;
  }};
  const setVal = (el, val) => {{
    const proto = Object.getPrototypeOf(el);
    const desc = Object.getOwnPropertyDescriptor(proto, 'value');
    if (desc && desc.set) desc.set.call(el, val); else el.value = val;
    el.dispatchEvent(new Event('input', {{bubbles:true}}));
    el.dispatchEvent(new Event('change', {{bubbles:true}}));
  }};
  let el = document.elementFromPoint(x, y);
  if (!el) return {{ok:false, message:'未找到坐标元素'}};
  const target = el.closest('button,[role="button"],input,textarea,.ant-select,.ant-picker,.ant-checkbox-wrapper,.ant-radio-wrapper,.ant-tabs-tab,th') || el;
  const tag = target.tagName.toLowerCase();
  const cls = target.className?.toString?.() || '';
  const type = target.getAttribute('type') || '';
  if (target.disabled || target.getAttribute('aria-disabled') === 'true' || cls.includes('disabled')) return {{ok:false, disabled:true, message:'控件禁用'}};
  target.scrollIntoView({{block:'center', inline:'center'}});
  await new Promise(r => setTimeout(r, 80));
  if ((tag === 'input' || tag === 'textarea') && !['checkbox','radio'].includes(type)) {{
    target.focus();
    setVal(target, 'TEST_NO_SAVE');
    await new Promise(r => setTimeout(r, 150));
    setVal(target, '');
    return {{ok:true, action:'input-clear'}};
  }}
  target.click();
  await new Promise(r => setTimeout(r, 500));
  return {{ok:true, action:'click'}};
}})()
"""
    try:
        return cdp.eval(script, timeout=8) or {"ok": True}
    except Exception as exc:
        return {"ok": False, "message": str(exc)}


def overlay_text(cdp: CDP) -> str:
    script = r"""
(() => {
  const visible = el => {
    if (!el) return false;
    const s = getComputedStyle(el);
    const r = el.getBoundingClientRect();
    return s.display !== 'none' && s.visibility !== 'hidden' && r.width > 0 && r.height > 0;
  };
  const selectors = [
    '.ant-modal',
    '.ant-drawer',
    '.ant-select-dropdown',
    '.ant-picker-dropdown',
    '.ant-dropdown',
    '.ant-message',
    '.ant-popover'
  ].join(',');
  return [...document.querySelectorAll(selectors)]
    .filter(visible)
    .map(el => (el.innerText || el.textContent || '').replace(/\s+/g,' ').trim())
    .filter(Boolean)
    .join('；');
})()
"""
    try:
        return clipped(cdp.eval(script, timeout=5), 1200)
    except Exception:
        return ""


def prompt_text(cdp: CDP) -> str:
    script = r"""
(() => {
  const visible = el => {
    if (!el) return false;
    const s = getComputedStyle(el);
    const r = el.getBoundingClientRect();
    return s.display !== 'none' && s.visibility !== 'hidden' && r.width > 0 && r.height > 0;
  };
  const selectors = [
    '.ant-message',
    '.ant-notification',
    '.ant-alert',
    '.ant-modal',
    '.ant-drawer',
    '.ant-popconfirm',
    '[role="alert"]',
    '[role="dialog"]'
  ].join(',');
  const seen = new Set();
  const prompts = [...document.querySelectorAll(selectors)]
    .filter(visible)
    .map(el => (el.innerText || el.textContent || '').replace(/\s+/g,' ').trim())
    .filter(Boolean)
    .filter(txt => {
      if (seen.has(txt)) return false;
      seen.add(txt);
      return true;
    });
  const bodyText = (document.body?.innerText || '').replace(/\s+/g,' ').trim();
  if (prompts.length === 0 && bodyText.length > 0 && bodyText.length < 1200 &&
      /错误|异常|失败|无权限|登录失效|重新登录|请求超时|服务器|Error|Exception|Network/i.test(bodyText)) {
    prompts.push(bodyText.slice(0, 600));
  }
  return prompts.join('；');
})()
"""
    try:
        return clipped(cdp.eval(script, timeout=5), 1200)
    except Exception:
        return ""


def abnormal_prompt(prompt: str, status: str) -> bool:
    source = f"{prompt} {status}"
    return any(
        marker in source
        for marker in [
            "异常",
            "错误",
            "失败",
            "报错",
            "无权限",
            "权限",
            "未授权",
            "登录失效",
            "重新登录",
            "请求失败",
            "接口",
            "服务器",
            "超时",
            "Error",
            "Exception",
            "Network",
        ]
    )


def close_overlays(cdp: CDP) -> None:
    script = r"""
(async () => {
  const visible = el => {
    if (!el) return false;
    const s = getComputedStyle(el);
    const r = el.getBoundingClientRect();
    return s.display !== 'none' && s.visibility !== 'hidden' && r.width > 0 && r.height > 0;
  };
  const safeText = /取消|关闭|关 闭|返回|取 消/;
  const buttons = [...document.querySelectorAll('.ant-modal button,.ant-drawer button,.ant-modal-close,.ant-drawer-close')]
    .filter(visible);
  const safe = buttons.find(b => safeText.test((b.innerText || b.textContent || b.title || '').replace(/\s+/g,''))) ||
    buttons.find(b => (b.className || '').toString().includes('close'));
  if (safe) {
    safe.click();
    await new Promise(r => setTimeout(r, 250));
  }
  document.dispatchEvent(new KeyboardEvent('keydown', {key:'Escape', code:'Escape', bubbles:true}));
  document.dispatchEvent(new KeyboardEvent('keyup', {key:'Escape', code:'Escape', bubbles:true}));
  return true;
})()
"""
    try:
        cdp.eval(script, timeout=5)
    except Exception:
        pass
    try:
        cdp.esc()
    except Exception:
        pass


def expected(control: dict) -> str:
    kind = text(control.get("kind"))
    if "下拉" in kind or "选择器" in kind:
        return "点击后下拉面板可展开，选项内容可识别，页面保持可继续操作。"
    if "日期" in kind:
        return "点击后日期/时间面板可打开，面板内容可识别，关闭后页面保持可继续操作。"
    if "输入框" in kind:
        return "输入测试文本后可清空，未触发保存类业务动作。"
    if "弹窗" in kind:
        return "点击后弹窗或抽屉可打开，关闭后返回原页面。"
    if "表头" in kind:
        return "表头字段可见；如支持排序，点击后页面保持可继续操作。"
    return "点击后控件响应符合预期，页面无阻塞异常。"


def add_record(records: list[dict], module: str, page: dict, control: dict, status: str, detail: str, prompt: str, clicked: str, risk: str) -> None:
    records.append(
        {
            "序号": len(records) + 1,
            "模块": module,
            "菜单路径": page.get("fullPathText", ""),
            "所属页面": page.get("name", ""),
            "路由地址": page.get("path", ""),
            "区域": control.get("area", ""),
            "按钮/控件": control.get("name", ""),
            "控件类型": control.get("kind", ""),
            "点击前置": "页面已打开，控件在当前视口或滚动后可见",
            "预期结果": expected(control),
            "实测状态": status,
            "展开/弹窗/选项内容": detail,
            "异常/弹窗提示": prompt,
            "是否执行点击": clicked,
            "风险说明": risk,
        }
    )


def audit_page(cdp: CDP, base_url: str, module: str, page: dict, max_controls: int, safe_mode: bool) -> tuple[list[dict], int, int]:
    cdp.navigate(page_url(base_url, page["path"]))
    wait_ready(cdp, 2.5)
    install_helper(cdp)
    expand_filters(cdp)
    wait_ready(cdp, 0.7)
    install_helper(cdp)
    controls = collect_controls(cdp)
    controls = controls[:max_controls]

    records: list[dict] = []
    click_count = 0
    risk_skip_count = 0
    for control in controls:
        name = text(control.get("name")) or text(control.get("kind")) or "控件"
        control["name"] = clipped(name, 120)
        if control.get("disabled"):
            add_record(records, module, page, control, "禁用未点击", "", "", "否", "")
            continue
        if safe_mode and is_risky(name):
            risk_skip_count += 1
            add_record(records, module, page, control, "风险动作未点击", "", "", "否", "生产环境安全边界：该动作可能改变业务数据或触发导出/下载。")
            continue

        result = interact(cdp, control)
        click_count += 1 if result.get("ok") else 0
        time.sleep(0.25)
        detail = overlay_text(cdp)
        prompt = prompt_text(cdp)
        if result.get("disabled"):
            status = "禁用未点击"
            clicked = "否"
        elif not result.get("ok"):
            status = "点击失败或未响应"
            clicked = "否"
            detail = detail or clipped(result.get("message"), 300)
            prompt = prompt or detail
        elif result.get("action") == "input-clear":
            status = "已点击/输入/清空"
            clicked = "是"
        elif detail and ("modal" in text(control.get("className")).lower() or "弹窗" in detail or "确定" in detail or "取消" in detail):
            status = "已点击，弹窗/抽屉已打开并关闭"
            clicked = "是"
        elif detail:
            status = "已打开"
            clicked = "是"
        else:
            status = "已点击"
            clicked = "是"
        if abnormal_prompt(prompt, status):
            status = "点击后出现异常提示"
        add_record(records, module, page, control, status, detail, prompt, clicked, "")
        close_overlays(cdp)
        time.sleep(0.15)
    return records, click_count, risk_skip_count


def style_workbook(wb: Workbook) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for ws in wb.worksheets:
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.border = border
        ws.freeze_panes = "A2"
        for idx, col in enumerate(ws.columns, start=1):
            values = [text(c.value) for c in col[:80]]
            width = min(max([len(v) for v in values] + [10]) + 2, 55)
            ws.column_dimensions[get_column_letter(idx)].width = width


def write_excel(records: list[dict], pages: list[dict], page_stats: dict[str, dict], output_path: Path, module: str, safe_mode: bool) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "按钮点击测试表"
    ws.append(OUTPUT_HEADERS)
    for record in records:
        ws.append([record.get(h, "") for h in OUTPUT_HEADERS])
    if len(records) > 0:
        end_col = get_column_letter(len(OUTPUT_HEADERS))
        tab = Table(displayName="KMMOMUiClickRecords", ref=f"A1:{end_col}{len(records)+1}")
        tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        ws.add_table(tab)

    ws2 = wb.create_sheet("页面汇总")
    ws2.append(["序号", "页面", "菜单路径", "路由地址", "测试记录数", "已执行点击数", "风险跳过数", "识别未点击数"])
    for i, page in enumerate(pages, start=1):
        stat = page_stats.get(page["path"], {})
        ws2.append(
            [
                i,
                page.get("name", ""),
                page.get("fullPathText", ""),
                page.get("path", ""),
                stat.get("records", 0),
                stat.get("clicks", 0),
                stat.get("risk_skips", 0),
                stat.get("not_clicked", 0),
            ]
        )

    ws3 = wb.create_sheet("测试说明")
    notes = [
        ("KMMOM UI 点击测试", ""),
        ("模块", module),
        ("生成时间", dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("安全模式", "是" if safe_mode else "否"),
        ("安全边界", "默认不执行保存、提交、确定、删除、导出下载、初始化、发布、下发、审批通过、驳回、发送消息等可能产生业务副作用的最终动作。"),
        ("异常/弹窗提示记录", "点击过程中如出现 toast、message、notification、alert、弹窗、抽屉、权限提示、登录失效、接口失败、空白页或异常页，必须记录可见提示原文和是否阻断操作。"),
        ("说明", "本文件由 kmmom-ui-test-generator skill 自动生成，用于页面按钮/控件点击测试记录和后续 MeterSphere 用例转换。"),
    ]
    for row in notes:
        ws3.append(row)

    style_workbook(wb)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Audit KMMOM module UI controls through a controllable Chrome CDP session.")
    parser.add_argument("--base-url", default=DEFAULT_BASE_URL)
    parser.add_argument("--debug-url", default=DEFAULT_DEBUG_URL)
    parser.add_argument("--workspace", default=os.getcwd())
    parser.add_argument("--module", default="制造执行")
    parser.add_argument("--output-dir", default="")
    parser.add_argument("--date", default=dt.datetime.now().strftime("%Y%m%d"))
    parser.add_argument("--username", default=os.environ.get("KMMOM_USERNAME", ""))
    parser.add_argument("--password-env", default="KMMOM_PASSWORD")
    parser.add_argument("--safe-mode", action="store_true", default=False)
    parser.add_argument("--unsafe-allow-final-actions", action="store_true", default=False)
    parser.add_argument("--list-pages-only", action="store_true")
    parser.add_argument("--max-controls-per-page", type=int, default=260)
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    workspace = Path(args.workspace)
    output_dir = Path(args.output_dir) if args.output_dir else workspace / "outputs" / "04_测试执行" / "02_执行记录"
    safe_mode = args.safe_mode or not args.unsafe_allow_final_actions
    password = os.environ.get(args.password_env, "")

    cdp = connect(args.debug_url, args.base_url)
    cdp.navigate(args.base_url)
    wait_ready(cdp, 2)
    login_if_needed(cdp, args.username, password)

    pages = get_module_pages(cdp, args.module)
    if not pages:
        raise RuntimeError(f"No leaf pages found for module: {args.module}. Confirm login state and menu permissions.")
    print(f"{args.module}页面数：{len(pages)}")
    for i, page in enumerate(pages, start=1):
        print(f"[{i}/{len(pages)}] {page.get('fullPathText')} -> {page.get('path')}")
    if args.list_pages_only:
        return

    all_records: list[dict] = []
    page_stats: dict[str, dict] = {}
    for i, page in enumerate(pages, start=1):
        print(f"[{i}/{len(pages)}] 开始：{page.get('fullPathText')}")
        records, clicks, risk_skips = audit_page(cdp, args.base_url, args.module, page, args.max_controls_per_page, safe_mode)
        all_records.extend(records)
        page_stats[page["path"]] = {
            "records": len(records),
            "clicks": clicks,
            "risk_skips": risk_skips,
            "not_clicked": sum(1 for r in records if r.get("是否执行点击") != "是"),
        }
        print(f"[{i}/{len(pages)}] 完成：记录 {len(records)}，点击 {clicks}，风险跳过 {risk_skips}")

    output_name = f"KMMOM-{args.module}模块-界面按钮点击测试表-{args.date}.xlsx"
    output_path = output_dir / output_name
    write_excel(all_records, pages, page_stats, output_path, args.module, safe_mode)
    print(output_path.resolve())
    print(f"records={len(all_records)}")


if __name__ == "__main__":
    main()
