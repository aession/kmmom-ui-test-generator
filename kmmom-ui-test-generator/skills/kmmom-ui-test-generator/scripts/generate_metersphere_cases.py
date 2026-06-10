# -*- coding: utf-8 -*-
from __future__ import annotations

import argparse
import datetime as dt
import os
import re
from collections import Counter
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


CASE_HEADERS = [
    "用例名称",
    "所属模块",
    "标签",
    "前置条件",
    "步骤描述",
    "预期结果",
    "编辑模式",
    "备注",
    "用例状态",
    "责任人",
    "用例等级",
]


def text(value) -> str:
    if value is None:
        return ""
    return str(value).replace("\r\n", "\n").replace("\r", "\n").strip()


def shorten(value, limit: int = 1200) -> str:
    value = text(value)
    if len(value) <= limit:
        return value
    return value[: limit - 20] + "...[内容已截断]"


def infer_module(source: Path, fallback: str = "") -> str:
    if fallback:
        return fallback
    m = re.search(r"KMMOM-(.+?)模块-", source.name)
    return m.group(1) if m else "KMMOM"


def infer_date(source: Path, fallback: str = "") -> str:
    if fallback:
        return fallback
    m = re.search(r"(\d{8})", source.name)
    return m.group(1) if m else dt.datetime.now().strftime("%Y%m%d")


def module_path(menu_path: str, page: str) -> str:
    parts = [p.strip() for p in text(menu_path).split(">") if p.strip()]
    if not parts:
        parts = ["KMMOM", page or "页面按钮控件"]
    if parts[0] != "KMMOM":
        parts.insert(0, "KMMOM")
    return "/" + "/".join(parts)


def control_name(row: dict) -> str:
    name = text(row.get("按钮/控件"))
    kind = text(row.get("控件类型"))
    if not name or name == "控件":
        name = kind or "控件"
    return shorten(name, 80)


def action_word(row: dict) -> str:
    kind = text(row.get("控件类型"))
    status = text(row.get("实测状态"))
    clicked = text(row.get("是否执行点击"))
    if clicked != "是":
        return "识别"
    if "下拉" in kind or "选择器" in kind:
        return "展开/查看"
    if "日期" in kind:
        return "打开日期面板"
    if "输入框" in kind:
        return "输入并清空"
    if "复选框" in kind or "单选" in kind:
        return "点击/切换"
    if "弹窗" in status or "已打开" in status:
        return "点击打开"
    return "点击"


def priority(row: dict) -> str:
    area = text(row.get("区域"))
    kind = text(row.get("控件类型"))
    risk = text(row.get("风险说明"))
    if area == "公共浮动区":
        return "P3"
    if risk or "风险" in text(row.get("实测状态")):
        return "P1"
    if area in {"筛选区", "页面条件区", "工具栏"}:
        return "P1"
    if "按钮" in kind:
        return "P1"
    return "P2"


def scenario_type(row: dict) -> str:
    area = text(row.get("区域"))
    kind = text(row.get("控件类型"))
    risk = text(row.get("风险说明"))
    if risk:
        return "安全边界"
    if "筛选" in area or "条件" in area:
        return "查询条件"
    if "表格" in area:
        return "表格交互"
    if "弹窗" in area:
        return "弹窗控件"
    if "下拉" in kind or "选择器" in kind:
        return "下拉选项"
    if "公共浮动区" in area:
        return "辅助控件"
    return "界面交互"


def build_steps(row: dict) -> list[str]:
    page = text(row.get("所属页面"))
    route = text(row.get("路由地址"))
    area = text(row.get("区域"))
    name = control_name(row)
    pre = text(row.get("点击前置")) or "页面已打开"
    risk = text(row.get("风险说明"))
    clicked = text(row.get("是否执行点击"))
    action = action_word(row)
    steps = [
        f"进入【{page}】页面，页面路由为【{route}】。",
        f"在【{area}】区域定位【{name}】控件，并确认执行前置为：{pre}。",
    ]
    if clicked == "是":
        steps.append(f"对【{name}】执行【{action}】操作。")
    else:
        steps.append(f"按安全边界仅识别【{name}】控件，不执行最终点击；风险说明：{risk or '无'}。")
    steps.append("观察页面提示、下拉选项、弹窗/抽屉、控件状态或页面可用性，完成后关闭弹层返回页面。")
    return steps


def build_expected(row: dict) -> list[str]:
    page = text(row.get("所属页面"))
    route = text(row.get("路由地址"))
    area = text(row.get("区域"))
    kind = text(row.get("控件类型")) or "控件"
    name = control_name(row)
    clicked = text(row.get("是否执行点击"))
    risk = text(row.get("风险说明"))
    expected = text(row.get("预期结果")) or "控件响应符合页面交互预期，页面保持可继续操作。"
    actual = text(row.get("实测状态"))
    detail = shorten(row.get("展开/弹窗/选项内容"), 600)
    prompt = shorten(row.get("异常/弹窗提示"), 600)
    expects = [
        f"页面成功打开，当前所属页面为【{page}】，路由为【{route}】。",
        f"【{area}】区域可见【{name}】，控件类型识别为【{kind}】。",
    ]
    if clicked == "是":
        expects.append(expected)
    else:
        expects.append(f"控件未执行最终点击，记录为安全识别；风险说明：{risk or '无'}。")
    if detail:
        expects.append(f"展开/弹窗/选项内容可识别，参考内容：{detail}")
    if prompt:
        expects.append(f"异常/弹窗提示已记录，参考提示：{prompt}")
    if not detail and not prompt:
        expects.append(f"操作后状态可观察，参考实测状态：{actual or '页面无阻塞异常'}。")
    return expects


def build_case(row: dict, source: Path, module: str, prefix: str) -> dict:
    seq = int(row.get("序号") or 0)
    page = text(row.get("所属页面"))
    area = text(row.get("区域"))
    name = control_name(row)
    case_no = f"{prefix}-{seq:03d}" if prefix else f"MES-UI-{seq:03d}"
    case_name = shorten(f"{case_no} 验证【{page}】页面【{area}-{name}】{action_word(row)}", 180)
    remark_parts = [
        f"来源测试记录: {source.name} 第{seq}条",
        f"场景类型: {scenario_type(row)}",
        f"菜单路径: {text(row.get('菜单路径'))}",
        f"路由地址: {text(row.get('路由地址'))}",
        f"控件类型: {text(row.get('控件类型')) or '控件'}",
        f"是否执行点击: {text(row.get('是否执行点击'))}",
        f"历史实测状态: {text(row.get('实测状态'))}",
    ]
    if text(row.get("展开/弹窗/选项内容")):
        remark_parts.append(f"展开/弹窗/选项内容: {shorten(row.get('展开/弹窗/选项内容'), 600)}")
    if text(row.get("异常/弹窗提示")):
        remark_parts.append(f"异常/弹窗提示: {shorten(row.get('异常/弹窗提示'), 600)}")
    if text(row.get("风险说明")):
        remark_parts.append(f"风险说明: {text(row.get('风险说明'))}")

    preconditions = [
        "1. 使用具备目标模块权限的账号登录 KMMOM 系统。",
        f"2. 当前账号可访问菜单路径：{text(row.get('菜单路径'))}。",
        "3. 当前为页面按钮/控件点击测试；生产环境默认不执行保存、提交、确定、删除、导出下载、初始化、发布、下发等最终业务动作。",
    ]

    return {
        "用例名称": case_name,
        "所属模块": module_path(row.get("菜单路径"), page),
        "标签": f"KMMOM,{module},页面按钮控件点击",
        "前置条件": "\n".join(preconditions),
        "步骤": build_steps(row),
        "预期": build_expected(row),
        "编辑模式": "TEST",
        "备注": "\n".join(remark_parts),
        "用例状态": "未开始",
        "责任人": "",
        "用例等级": priority(row),
    }


def control_step(row: dict) -> str:
    area = text(row.get("区域"))
    name = control_name(row)
    clicked = text(row.get("是否执行点击"))
    risk = text(row.get("风险说明"))
    if clicked == "是":
        return f"验证【{area}】区域【{name}】控件，执行【{action_word(row)}】操作。"
    return f"验证【{area}】区域【{name}】控件，按安全边界仅识别不执行最终点击；风险说明：{risk or '无'}。"


def control_expected(row: dict) -> str:
    clicked = text(row.get("是否执行点击"))
    expected = text(row.get("预期结果"))
    actual = text(row.get("实测状态"))
    detail = shorten(row.get("展开/弹窗/选项内容"), 600)
    prompt = shorten(row.get("异常/弹窗提示"), 600)
    risk = text(row.get("风险说明"))
    parts = []
    if clicked == "是":
        parts.append(expected or "控件响应符合页面交互预期，页面保持可继续操作。")
    else:
        parts.append(f"控件可见性、状态和风险原因被记录；不执行最终业务动作。{risk or ''}")
    if detail:
        parts.append(f"参考展开/弹窗/选项内容：{detail}")
    if prompt:
        parts.append(f"参考异常/弹窗提示：{prompt}")
    if not detail and not prompt and actual:
        parts.append(f"参考实测状态：{actual}")
    return " ".join(p for p in parts if p).strip()


def build_page_case(page_rows: list[dict], source: Path, module: str, prefix: str, page_index: int) -> dict:
    first = page_rows[0]
    page = text(first.get("所属页面"))
    route = text(first.get("路由地址"))
    case_no = f"{prefix}-PAGE-{page_index:03d}" if prefix else f"MES-UI-PAGE-{page_index:03d}"
    case_name = shorten(f"{case_no} 验证【{page}】页面按钮/控件自动化点击巡检", 180)
    risk_count = sum(1 for r in page_rows if text(r.get("是否执行点击")) != "是")
    click_count = sum(1 for r in page_rows if text(r.get("是否执行点击")) == "是")
    areas = ", ".join(sorted({text(r.get("区域")) for r in page_rows if text(r.get("区域"))}))

    steps = [
        f"进入【{page}】页面，页面路由为【{route}】。",
        "展开页面筛选区、页面条件区或可展开区域，确认页面主要控件已加载。",
    ]
    expects = [
        f"页面成功打开，当前所属页面为【{page}】，路由为【{route}】。",
        f"页面控件区域可识别，覆盖区域包括：{areas or '页面可见区域'}。",
    ]
    for row in page_rows:
        steps.append(control_step(row))
        expects.append(control_expected(row))

    remark_parts = [
        f"来源测试记录: {source.name}",
        "场景类型: 页面级按钮/控件自动化巡检",
        f"菜单路径: {text(first.get('菜单路径'))}",
        f"路由地址: {route}",
        f"页面控件记录数: {len(page_rows)}",
        f"已执行点击/展开/输入记录数: {click_count}",
        f"安全识别或风险跳过记录数: {risk_count}",
        "说明: 本用例按页面聚合，页面内每个按钮/控件作为步骤展开；如需一控件一用例，可使用 --granularity control。",
    ]

    preconditions = [
        "1. 使用具备目标模块权限的账号登录 KMMOM 系统。",
        f"2. 当前账号可访问菜单路径：{text(first.get('菜单路径'))}。",
        "3. 当前为页面按钮/控件点击测试；生产环境默认不执行保存、提交、确定、删除、导出下载、初始化、发布、下发等最终业务动作。",
        "4. 页面级用例执行时，应按步骤逐项记录控件响应；若遇到业务写入类动作，仅验证可见性和风险说明。",
        "5. 若点击过程中出现 toast、弹窗、权限提示、接口失败或异常页，应记录提示原文并判断是否阻断后续操作。",
    ]

    return {
        "用例名称": case_name,
        "所属模块": module_path(first.get("菜单路径"), page),
        "标签": f"KMMOM,{module},页面按钮控件点击,页面级",
        "前置条件": "\n".join(preconditions),
        "步骤": steps,
        "预期": expects,
        "编辑模式": "TEST",
        "备注": "\n".join(remark_parts),
        "用例状态": "未开始",
        "责任人": "",
        "用例等级": "P1" if click_count or risk_count else "P2",
    }


def write_case(ws, case: dict) -> int:
    steps = case["步骤"]
    expects = case["预期"]
    count = max(len(steps), len(expects))
    for idx in range(count):
        if idx == 0:
            ws.append(
                [
                    case["用例名称"],
                    case["所属模块"],
                    case["标签"],
                    case["前置条件"],
                    steps[idx] if idx < len(steps) else "",
                    expects[idx] if idx < len(expects) else "",
                    case["编辑模式"],
                    case["备注"],
                    case["用例状态"],
                    case["责任人"],
                    case["用例等级"],
                ]
            )
        else:
            ws.append(["", "", "", "", steps[idx] if idx < len(steps) else "", expects[idx] if idx < len(expects) else "", "", "", "", "", ""])
    return count


def style(ws) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    widths = [46, 34, 28, 48, 56, 60, 12, 58, 12, 12, 12]
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border
    for i, width in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = width
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Convert KMMOM UI click records into MeterSphere import cases.")
    parser.add_argument("--source", required=True, help="Source KMMOM UI click record Excel.")
    parser.add_argument("--workspace", default=os.getcwd())
    parser.add_argument("--output-dir", default="")
    parser.add_argument("--module", default="")
    parser.add_argument("--date", default="")
    parser.add_argument("--prefix", default="MES-UI")
    parser.add_argument("--granularity", choices=["page", "control"], default="page", help="page: one case per page; control: one case per UI control.")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    source = Path(args.source)
    if not source.is_absolute():
        source = Path(args.workspace) / source
    if not source.exists():
        raise FileNotFoundError(source)

    module = infer_module(source, args.module)
    date_s = infer_date(source, args.date)
    output_dir = Path(args.output_dir) if args.output_dir else Path(args.workspace) / "outputs" / "03_测试用例" / "03_MeterSphere导入"
    output_dir.mkdir(parents=True, exist_ok=True)
    output = output_dir / f"KMMOM-{module}模块-页面按钮控件点击测试用例-MeterSphere导入-{date_s}.xlsx"

    wb_source = load_workbook(source, read_only=True, data_only=True)
    ws_source = wb_source.worksheets[0]
    headers = [c.value for c in next(ws_source.iter_rows(min_row=1, max_row=1))]
    rows = []
    for values in ws_source.iter_rows(min_row=2, values_only=True):
        row = dict(zip(headers, values))
        if text(row.get("所属页面")) and text(row.get("按钮/控件")):
            rows.append(row)
    wb_source.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "测试用例"
    ws.append(CASE_HEADERS)

    page_counter = Counter()
    level_counter = Counter()
    excel_rows = 0
    if args.granularity == "control":
        case_total = len(rows)
        for row in rows:
            case = build_case(row, source, module, args.prefix)
            excel_rows += write_case(ws, case)
            page_counter[text(row.get("所属页面"))] += 1
            level_counter[case["用例等级"]] += 1
    else:
        grouped: dict[tuple[str, str], list[dict]] = {}
        order: list[tuple[str, str]] = []
        for row in rows:
            key = (text(row.get("所属页面")), text(row.get("路由地址")))
            if key not in grouped:
                grouped[key] = []
                order.append(key)
            grouped[key].append(row)
        case_total = len(order)
        for index, key in enumerate(order, start=1):
            page_rows = grouped[key]
            case = build_page_case(page_rows, source, module, args.prefix, index)
            excel_rows += write_case(ws, case)
            page_counter[key[0]] += len(page_rows)
            level_counter[case["用例等级"]] += 1

    style(ws)
    wb.save(output)
    print(output.resolve())
    print(f"granularity={args.granularity}")
    print(f"cases={case_total}")
    print(f"excel_rows={excel_rows + 1}")
    print("pages=" + "; ".join(f"{k}:{v}" for k, v in page_counter.items()))
    print("levels=" + "; ".join(f"{k}:{v}" for k, v in sorted(level_counter.items())))


if __name__ == "__main__":
    main()
